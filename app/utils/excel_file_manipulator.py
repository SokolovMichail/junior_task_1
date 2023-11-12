from io import BytesIO
from typing import Optional

import openpyxl
import pandas as pd
from sqlalchemy.orm import Session

from app.schemes.file_scheme import FileModel
from app.utils.dataframe_processor import DataframeProcessor
from app.utils.model_utils.file_model_utils import FileModelUtils


class ExcelFileManipulator:

    @staticmethod
    def preprocess_pandas_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        renaming_index = {df.columns[0]: "Код", df.columns[1]: "Наименование проекта"}
        for i in range(2, len(df.columns), 2):
            renaming_index[df.columns[i]] = str(df.columns[i]) + "_plan"
            renaming_index[df.columns[i + 1]] = str(df.columns[i]) + "_fact"
        result = df.rename(columns=renaming_index).tail(-1)
        return result

    @staticmethod
    def load_values_into_db(df: pd.DataFrame, filename: str, db: Session):
        preprocessed_dataframe = ExcelFileManipulator.preprocess_pandas_dataframe(df)
        file = FileModelUtils.db_create_new_file(db, filename)
        DataframeProcessor.process_dataframe(db, preprocessed_dataframe, file)
        return file

    @staticmethod
    def return_file_from_db(db: Session, requested_file: FileModel) -> Optional[BytesIO]:
        result_dataframe = DataframeProcessor.construct_result_dataframe(requested_file, db)
        if len(result_dataframe) == 0:
            return None
        binary_file = BytesIO()
        result_dataframe.to_excel(binary_file,sheet_name='data', startrow=1,index=False)
        ExcelFileManipulator.postprocess_dumped_dataframe(binary_file,result_dataframe.columns)
        return binary_file


    @staticmethod
    def postprocess_dumped_dataframe(excel_file: BytesIO, columns):
        workbook = openpyxl.load_workbook(excel_file)
        ws = workbook.active
        for i in range(3,len(columns)+1,2):
            ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i + 1)
            ws.cell(row=1,column=i).value = columns[i].split("_")[0]
            ws.cell(row=2,column=i).value = "план"
            ws.cell(row=2, column=i+1).value = "факт"
            #ws[1][i].value = datetime.strptime(columns[i].split("_")[0], "%d.%m.%Y")
        workbook.save(excel_file)

